# {input_path} source input excel path
# {output_path} output ABSOLUTE directory
# {output_filename}: output file name. If it's none, it should be [input_filename].[output_format]
# {overwrite}: If the output filename exists, overwrite it or not.
#   if not , then add a random suffix to the output filename
# {output_format} target file format, default is CSV file
# {simple_mode}, default is true, read the 1st or specified sheet and convert the whole sheet to {output_format} file
# {read_all}, it works only when simple_mode=True, then read all the sheets and ignore {sheet_name} param 
# {keep_header} default it's True. 
# {header}, default it's 0, means no header. if header>1, 1st row is header for each sheet.
# {sheet_name}, it works only {when} simple_mode = True. Then only read the specified sheet
# {merge}, merge the data in all {she}ets into a file or not, default is True
# {convertion_func_list},用于过滤黑白单之类的操作
# {**sheet}, parameters for {eac}h sheet if it has, the struct should be:
# sheet ={
#    "format_parameters" : {}
#     [sheet_name_1]:{
#         start_row: xx,
#         ignore_end_row: xx,
#         start_column: xx,
#         column_width:xx
#     },
#     [sheet_name_2]:{
#             start_row: xx,
#             ignore_end_row: xx,
#             start_column: xx,
#             column_width:xx
#         }
# }
import logging
import openpyxl
import os, csv

#filter out the empty row
def filter_empty_row( row:list , *args ):
    if all( [ v is None or str(v).strip(" ")=='' for v in row] ):
        return None
    else :
        return row

class ExcelConverter:
    def __init__(self, keep_empty = False) -> None:
        self.convertion_func_list = []
        self.keep_empty =  keep_empty
    
        if not self.keep_empty:
           self.convertion_func_list.append(filter_empty_row) 


    def register_fun_list(self, fn_list):
        if fn_list is None or len(fn_list) ==0:
            return
        for fn in fn_list:
            self.register_convertion_fun(fn)

    def register_convertion_fun(self, fn):
        self.convertion_func_list.insert(0, fn)

    def clear_convertion_fun(self):
        self.convertion_func_list.clear()
        if not self.keep_empty:
           self.convertion_func_list.append(filter_empty_row) 
    
    def convert_xls2csv(self, input_path,
                        output_path = None,
                        output_filename = None,
                        overwrite = True,
                        output_format='csv',
                        output_encoding = 'utf8',
                        simple_mode = True,
                        read_all = False,
                        keep_header = True,
                        header = 0,
                        sheet_name = None,
                        merge= True,
                        **sheets_param):
        input_path = os.path.abspath(input_path)
        input_absoult_dir, input_ext = os.path.splitext(input_path)                     
        if not os.path.exists(input_path)\
                or input_ext not in (".xls", ".xlsx"):
            raise FileNotFoundError("Input excel file does not exists: " + input_path)
        
        if not simple_mode and (sheets_param is None or len(sheets_param) == 0):
            raise ValueError("Invalid parameter: simple_mode = False and sheets_params is none")
        
        file_format_parameters = sheets_param["_format_parameters"] if "_format_parameters" in sheets_param else None
        if file_format_parameters is not None and  type(file_format_parameters) is not dict:
            raise ValueError("format_parameters must be dict")

        if self.convertion_func_list is not None:
            for fn in self.convertion_func_list:
                if not callable(fn):
                    raise ValueError("convertion_func_list must be callable")

        output_path_dir =  output_path if output_path is not None else os.path.dirname(input_absoult_dir) 
        if not os.path.isdir(output_path_dir):
            os.makedirs(output_path_dir)
        
        base_filename = output_filename if output_filename is not None else os.path.basename(input_absoult_dir) 
        output_abs_basename = os.path.join(output_path_dir, base_filename)

        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
        # 整理需要处理的sheets 
        if simple_mode and read_all:
            sheets_name_list = wb.sheetnames
        elif not simple_mode and sheets_param is not None:
            sheets_name_list = list(sheets_param.keys())
            if "_format_parameters" in sheets_name_list:
                sheets_name_list.remove("_format_parameters")
        else:
            sheets_name_list = [wb.sheetnames[0]] if sheet_name is None else [sheet_name]

        sheets_name_list = list(set(sheets_name_list) & set(wb.sheetnames))
        # if donot overwrite the exists, then raise exception once the target file exists
        # output_file_path = ouput_path + output_name + [sheet name] + output_format
        if len(sheets_name_list) <= 1 or merge:
            output_abs_filename = output_abs_basename + "." + output_format
            flag_target_path_ready = overwrite and not os.path.exists(output_abs_filename)
        else:
            for sn in sheets_name_list:
                output_abs_filename = output_abs_basename + "_"+sn+"." + output_format
                flag_target_path_ready = overwrite and not os.path.exists(output_abs_filename)
                if not flag_target_path_ready:
                    break
        if not overwrite and not flag_target_path_ready:
            raise FileExistsError("Output file pathe exists: " + output_abs_filename)

        # 读取excel， 并转换成csv   
        if merge or len(sheets_name_list) <= 1:
            output_abs_filename = output_abs_basename + "." + output_format
        else:
            output_abs_filename = output_abs_basename + "_" + sheets_name_list[0] + "." + output_format

        output_abs_file = []
        fd, fdwriter = self.get_writer_by_format(output_format, output_abs_filename, encoding = output_encoding, file_format_parameters = file_format_parameters)
        output_abs_file.append(output_abs_filename)
        sheet_count = 0
        for sheet in sheets_name_list:
            if sheet_count > 0  and not merge:
                output_abs_filename = output_abs_basename + "_" + sheet + "." + output_format
                fd, fdwriter = self.get_writer_by_format(output_format, output_abs_filename, encoding = output_encoding, file_format_parameters = file_format_parameters)
                output_abs_file.append(output_abs_filename)

            read_param = dict()
            if simple_mode and len(sheets_name_list) > 1:
                start_row = 0
                if merge: 
                    if keep_header and sheet_count == 0:
                        #合并多个sheet 且保留header， 只有sheet1, 保留header
                        start_row =  0  if sheet_count > 1  and header > 0 else header - 1
                    else: 
                        #合并多个sheet 且不保留header或第二sheet， 所有sheet不保留header
                        start_row =  header if header > 0 else 0
                else:
                    if keep_header:
                        #不合并多个sheet且保留header， 全部保留header
                        start_row =  header-1 if header > 0 else 0 
                    else: 
                        #不合并多个sheet且不保留header， 所有sheet不保留header
                        start_row =  header if header > 0 else 0

                read_param["start_row"] = start_row
                sheets_param[sheet] = read_param

            sheet_count += 1
            
            if self.convertion_func_list is None:
                for row in self.read_sheet(wb[sheet], sheet, **sheets_param):
                    # 如果要支持其他文件格式， 则需要重装对应文件的writer，使其支持 writerow(iteror)
                    newline = [ '' if v.value is None else str(v.value) for v in row ]
                    fdwriter.writerow( newline )
            else:
                for row in self.read_sheet(wb[sheet], sheet, **sheets_param):
                    newline = [ '' if v.value is None else str(v.value) for v in row ]
                    for fn in self.convertion_func_list:
                        newline = fn(newline)

                    if newline is None:
                        continue
                    # 如果要支持其他文件格式， 则需要重装对应文件的writer，使其支持 writerow(iteror)
                    
                    newline = fn(newline, input_path, sheet)

            if not merge:
                fd.flush()
                fd.close()
        else:
            try: 
                fd.flush()
                fd.close()
            except:
                pass

        return output_abs_file
    # kwargs={
    # start_row: xx,
    # ignore_end_row: xx,
    # start_column: xx,
    # column_width:xx
    # }
    def read_sheet(self, worksheet, sheetname, **kwargs):
        start_row = 0 if sheetname not in kwargs or \
                "start_row" not in kwargs[sheetname] else kwargs[sheetname]["start_row"]
        ignore_end_row = 0 if sheetname not in kwargs or \
                "ignore_end_row" not in kwargs[sheetname] else kwargs[sheetname]["ignore_end_row"]
        start_column = 0 if sheetname not in kwargs or \
                "start_column" not in kwargs[sheetname] else kwargs[sheetname]["start_column"]
        row_count = worksheet.max_row
        col_count = worksheet.max_column
        column_width = col_count if sheetname not in kwargs or \
                "column_width" not in kwargs[sheetname] else kwargs[sheetname]["column_width"]
        end_column = start_column + column_width
        if start_column + ignore_end_row >= row_count:
            logging.warning( "start_row + ignore_end_row > total row count")
            return
        if start_column >= col_count:
            logging.warning( "The sheet has {0} columns, but start from {1} column".format(col_count, start_column))
            return

        offset = 0
        for row in worksheet.rows:
            if offset < start_row:
                offset += 1
                continue
            if offset + ignore_end_row > row_count:
                break
        
            yield  row[start_column: end_column]
            offset += 1



    def get_writer_by_format(self, output_format, output_abs_path, encoding = 'utf8', file_format_parameters = None):
        if output_format.lower() == 'csv':
            delimiter = file_format_parameters["delimiter"]  if file_format_parameters is not None and "delimiter" in file_format_parameters else ","
            doublequote = file_format_parameters["doublequote"]  if file_format_parameters is not None and  "doublequote" in file_format_parameters else True
            escapechar = file_format_parameters["escapechar"]  if file_format_parameters is not None and  "escapechar" in file_format_parameters else '"'
            lineterminator = file_format_parameters["lineterminator"]  if file_format_parameters is not None and  "lineterminator" in file_format_parameters else "\n"
            quotechar = file_format_parameters["quotechar"]  if file_format_parameters is not None and  "quotechar" in file_format_parameters else '"'
            quoting = file_format_parameters["quoting"]  if file_format_parameters is not None and  "quoting" in file_format_parameters else csv.QUOTE_MINIMAL
            skipinitialspace = file_format_parameters["skipinitialspace"]  if file_format_parameters is not None and  "skipinitialspace" in file_format_parameters else False
            strict = file_format_parameters["strict"]  if file_format_parameters is not None and  "strict" in file_format_parameters else False

            csv.register_dialect('excel2csv', 
                delimiter=delimiter , 
                doublequote = doublequote,
                escapechar = escapechar,
                lineterminator = lineterminator,
                quotechar = quotechar,
                quoting=quoting, 
                skipinitialspace = skipinitialspace,
                strict = strict
            )
            csvf = open( output_abs_path,'w', encoding=encoding, newline='')
            spanwriter=csv.writer(csvf, dialect='excel2csv')  
            return (csvf, spanwriter)


from datetime import datetime 
if __name__ == "__main__":
    pass
###############################################################
    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", overwrite = True)
    # print(datetime.now() - start_time )
###############################################################
    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", overwrite = True, header=1, read_all= True, merge=False)
    # print(datetime.now() - start_time )
###############################################################
    # sheet ={
    # "Sheet1":{
    #     'start_row': 0 ,
    #     'ignore_end_row': 2,
    #     'start_column': 1,
    #     'column_width': 5
    # },
    # "Sheet2":{
    #         # 'start_row': 0,
    #         # 'ignore_end_row': xx,
    #         'start_column': 1,
    #        'column_width': 5
    #     }
    # }
    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", 
    #     overwrite = True, 
    #     simple_mode=False,
    #     header=1, 
    #     read_all= True, 
    #     merge=False,
    #     **sheet)
    # print(datetime.now() - start_time )

    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", 
    #     overwrite = True, 
    #     simple_mode=False,
    #     header=1, 
    #     read_all= True, 
    #     merge=True,
    #     **sheet)
    # print(datetime.now() - start_time )

    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", 
    #     overwrite = True, 
    #     simple_mode=False,
    #     header=1, 
    #     read_all= True, 
    #     merge=True,
    #     **sheet)
    # print(datetime.now() - start_time )
###############################################################

    # test_list = [None, '','']
    # print(filter_empty_row(test_list))

###############################################################
    # sheet ={
    #     "_format_parameters" : {
    #         "quoting": csv.QUOTE_ALL,
    #         "delimiter":"\t"
    #     },
    #     "Sheet1":{
    #         'start_row': 0 ,
    #         'ignore_end_row': 2,
    #         # 'start_column': 1,
    #         'column_width': 5
    #     },
    #     "Sheet2":{
    #             # 'start_row': 0,
    #             # 'ignore_end_row': xx,
    #             # 'start_column': 1,
    #         'column_width': 5
    #         }
    #     }
    # print("start to converting")
    # start_time = datetime.now()
    # print(start_time)
    # excle2csv = ExcelConverter()
    # excle2csv.convert_xls2csv(r"C:\workspace\project\LEGO\code\CDP-CR\POC\calendar.xlsx", 
    #     overwrite = True, 
    #     simple_mode=False,
    #     header=1, 
    #     read_all= True, 
    #     merge=True,
    #     **sheet)
    # print(datetime.now() - start_time )